from enums.filter_operations import PhaseType

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import pandas as pd
import os


def create_scenario_xlsx(output: str, name: str, filtered_data_dict: dict[str, pd.DataFrame]) -> None:
    scenario_dict = {'Corrente': [], 'Cenário': []}
    for current, df in filtered_data_dict.items():
        scenario_dict['Corrente'].append(current)
        scenario_dict['Cenário'].append(df.iloc[1, 0])

    scenario_df = pd.DataFrame(scenario_dict)
    file_path = os.path.join(output, "scenario_names.xlsx")

    if os.path.exists(file_path):
        with pd.ExcelWriter(file_path, mode="a", if_sheet_exists="replace", engine="openpyxl") as writer:
            scenario_df.to_excel(writer, sheet_name=name, index=False)
    else:
        with pd.ExcelWriter(file_path, mode="w", engine="openpyxl") as writer:
            scenario_df.to_excel(writer, sheet_name=name, index=False)


def add_composition_value(composition_base: list[str], composition: pd.DataFrame, index: int, sheet: any) -> None:
    indices_in_base = [composition_base.index(item) for item in composition.columns]
    missing_items_indices = [i for i, item in enumerate(composition_base) if item not in composition.columns]

    for comp_value, comp_index in zip(composition.iloc[1, :].to_list(), indices_in_base):
        sheet.cell(row=7 + comp_index, column=4 + index, value=comp_value)

    for missing_indices in missing_items_indices:
        sheet.cell(row=7 + missing_indices, column=4 + index, value=0)


def create_final_composition_file(fraction_phase: PhaseType, name: str, fracao_final_output: str, filtered_data_dict: dict[str, pd.DataFrame]) -> None:
    current_dir = os.path.dirname(__file__) 
    base_fraction_file = os.path.abspath(os.path.join(current_dir, f'../../files/utils/BASE_FRAC_MOLAR_{fraction_phase.value}.xlsx'))

    wb = openpyxl.load_workbook(base_fraction_file)
    sheet = wb.active

    index = 0
    for corrente, df in filtered_data_dict.items():
        sheet.cell(row=4, column=4 + index, value=corrente)

        composition = df.iloc[:, 33:]
        composition_base = [sheet.cell(row=7 + row_index_value, column= 2).value for row_index_value in range(36)]

        columns_not_in_base = [col for col in composition.columns if col.lower() not in [c.lower() for c in composition_base]]
        base_not_in_columns = [col for col in composition_base if col.lower() not in [c.lower() for c in composition.columns]]

        if columns_not_in_base:
            indices_to_replace = [composition_base.index(item) for item in base_not_in_columns]

            for i, new_value in zip(indices_to_replace, columns_not_in_base):
                composition_base[i] = new_value
                sheet.cell(row=7 + i, column= 2, value=new_value)

        add_composition_value(composition_base, composition, index, sheet)

        spec_list = [4, 3, 15, 16, 2, 1, 30, 29, 14, 13, 5, 18, 17, 11, 10, 27, 26, 21, 20, 12, 28, 22, 7, 8]
        for row_index, spec_index in enumerate(spec_list):
            target_cell = sheet.cell(row=45 + row_index, column=4 + index)
            target_cell.value = df.iloc[1, spec_index]
            target_cell.alignment = Alignment(horizontal='center', vertical='center')

            if not target_cell.fill.patternType:
                fill = PatternFill(start_color="FFB4C6E7", end_color="FFB4C6E7", fill_type="solid")
                target_cell.fill = fill

        index += 3

    wb.save(os.path.join(fracao_final_output, f'{name}.xlsx'))


def save_excel_file(output_file: str, name: str, fraction_phase: PhaseType, filtered_data_dict: dict[str, pd.DataFrame]) -> None:
    output = os.path.join(output_file, fraction_phase.name)
    composicoes_output = os.path.join(output, 'Composicoes')
    fracao_final_output = os.path.join(output, 'Fracao_Final')

    if not os.path.exists(output):
        os.makedirs(output)

    if not os.path.exists(composicoes_output):
        os.makedirs(composicoes_output)

    if not os.path.exists(fracao_final_output):
        os.makedirs(fracao_final_output)

    create_scenario_xlsx(output, name, filtered_data_dict)

    composicoes_output_final = os.path.join(composicoes_output, f'composicoes_{name}.xlsx')

    writer = pd.ExcelWriter(composicoes_output_final)

    for sheet_name, dataframe in filtered_data_dict.items():
        dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

    writer.close()

    create_final_composition_file(fraction_phase, name, fracao_final_output, filtered_data_dict)
